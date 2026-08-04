from __future__ import annotations

import csv
import json
import math
import os
import re
import shutil
import statistics
import threading
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover
    Workbook = None
    WriteOnlyCell = None
    Alignment = None
    Border = None
    Font = None
    PatternFill = None
    Side = None


DATA_COLUMNS: tuple[str, ...] = (
    "timestamp_local",
    "elapsed_s",
    "interval_s",
    "current_a",
    "current_std_a",
    "current_sigma_a",
    "f_left_hz",
    "f_right_hz",
    "delta_f_hz",
    "delta_f_sigma_hz",
    "common_mode_hz",
    "valid_fraction",
    "all_samples_valid",
    "sample_count",
    "valid_sample_count",
    "left_state",
    "right_state",
    "invalid_reason",
    "relock_count",
    "lost_lock_count",
    "left_error_hz",
    "right_error_hz",
    "left_quality",
    "right_quality",
    "measured_update_rate_hz",
)

EXCEL_HEADERS: tuple[str, ...] = (
    "Timestamp (local)",
    "Elapsed (s)",
    "Aggregation interval (s)",
    "Current (A)",
    "Current std (A)",
    "Current uncertainty (A)",
    "Left resonance fL (Hz)",
    "Right resonance fR (Hz)",
    "Splitting Δf (Hz)",
    "Splitting uncertainty (Hz)",
    "Common-mode frequency (Hz)",
    "Valid fraction",
    "All samples valid",
    "Samples in interval",
    "Valid samples in interval",
    "Left peak state",
    "Right peak state",
    "Invalid reason",
    "Relock count",
    "Lost-lock count",
    "Left frequency error (Hz)",
    "Right frequency error (Hz)",
    "Left quality",
    "Right quality",
    "Measured update rate (Hz)",
)

NUMERIC_MEAN_FIELDS: tuple[str, ...] = (
    "estimated_current_a",
    "current_sigma_a",
    "left_frequency_hz",
    "right_frequency_hz",
    "splitting_hz",
    "delta_f_sigma_hz",
    "common_mode_hz",
    "left_error_hz",
    "right_error_hz",
    "left_quality",
    "right_quality",
)


def _finite_float(value: Any) -> float | None:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return numeric if math.isfinite(numeric) else None


def _mean(values: list[float]) -> float | None:
    return statistics.fmean(values) if values else None


def _sample_std(values: list[float]) -> float | None:
    return statistics.stdev(values) if len(values) >= 2 else (0.0 if values else None)


def _safe_filename(value: str, fallback: str) -> str:
    normalized = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff._-]+", "_", str(value).strip())
    normalized = normalized.strip("._-")
    return normalized[:80] or fallback


def _atomic_json_write(path: Path, payload: dict[str, Any]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    os.replace(temporary, path)


def _flatten_mapping(
    value: Any,
    prefix: str = "",
) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    if isinstance(value, dict):
        for key in sorted(value):
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(_flatten_mapping(value[key], child_prefix))
        return rows
    if isinstance(value, (list, tuple)):
        rows.append((prefix, json.dumps(value, ensure_ascii=False, default=str)))
        return rows
    rows.append((prefix, value))
    return rows


class CurrentTrackingRecorder:
    """1 秒级聚合并增量落盘；Excel 只是可下载视图，CSV 是恢复源。"""

    def __init__(
        self,
        *,
        base_dir: Path,
        interval_s: float,
        label: str,
        request_snapshot: dict[str, Any],
        device_snapshot: dict[str, Any],
    ) -> None:
        self.lock = threading.RLock()
        self.export_lock = threading.Lock()
        self.interval_s = max(0.1, float(interval_s))
        self.label = str(label).strip()
        now = datetime.now().astimezone()
        self.started_datetime = now
        self.session_id = f"{now:%Y%m%d_%H%M%S}_{uuid.uuid4().hex[:8]}"
        session_name = _safe_filename(
            f"{self.session_id}_{self.label}" if self.label else self.session_id,
            self.session_id,
        )
        self.session_dir = base_dir / session_name
        self.session_dir.mkdir(parents=True, exist_ok=False)
        self.csv_path = self.session_dir / "current_tracking_data.csv"
        self.metadata_path = self.session_dir / "metadata.json"
        self.xlsx_path = self.session_dir / f"current_tracking_{session_name}.xlsx"
        self.started_at = now.isoformat(timespec="milliseconds")
        self.ended_at: str | None = None
        self.status = "recording"
        self.request_snapshot = dict(request_snapshot)
        self.device_snapshot = dict(device_snapshot)
        self.rows_written = 0
        self.valid_rows = 0
        self.last_saved_at: str | None = None
        self.last_elapsed_s: float | None = None
        self.pending_points: list[dict[str, Any]] = []
        self.next_deadline_s: float | None = None
        self.closed = False
        self.last_export_error = ""
        self.error_message = ""
        self.failed_stage = ""
        self.error_code = ""
        self.error_hint = ""
        self._file = self.csv_path.open("w", newline="", encoding="utf-8-sig")
        self._writer = csv.DictWriter(self._file, fieldnames=DATA_COLUMNS)
        self._writer.writeheader()
        self._file.flush()
        self._write_metadata()

    def _metadata(self) -> dict[str, Any]:
        return {
            "session_id": self.session_id,
            "label": self.label,
            "status": self.status,
            "started_at": self.started_at,
            "ended_at": self.ended_at,
            "record_interval_s": self.interval_s,
            "rows_written": self.rows_written,
            "valid_rows": self.valid_rows,
            "last_saved_at": self.last_saved_at,
            "last_elapsed_s": self.last_elapsed_s,
            "csv_path": str(self.csv_path.resolve()),
            "xlsx_path": (
                str(self.xlsx_path.resolve()) if self.xlsx_path.exists() else None
            ),
            "last_export_error": self.last_export_error,
            "error_message": self.error_message,
            "failed_stage": self.failed_stage,
            "error_code": self.error_code,
            "error_hint": self.error_hint,
            "request": self.request_snapshot,
            "devices": self.device_snapshot,
        }

    def _write_metadata(self) -> None:
        _atomic_json_write(self.metadata_path, self._metadata())

    def status_dict(self) -> dict[str, Any]:
        with self.lock:
            payload = self._metadata()
            payload["download_available"] = self.rows_written > 0
            payload["csv_size_bytes"] = (
                self.csv_path.stat().st_size if self.csv_path.exists() else 0
            )
            payload["xlsx_size_bytes"] = (
                self.xlsx_path.stat().st_size if self.xlsx_path.exists() else 0
            )
            return payload

    def _aggregate(
        self,
        points: list[dict[str, Any]],
        interval_s: float,
    ) -> dict[str, Any]:
        if not points:
            raise ValueError("没有可聚合的跟踪点。")
        valid_points = [
            point
            for point in points
            if bool(point.get("valid"))
            and _finite_float(point.get("estimated_current_a")) is not None
        ]
        current_values = [
            value
            for point in valid_points
            if (value := _finite_float(point.get("estimated_current_a"))) is not None
        ]

        means: dict[str, float | None] = {}
        for field in NUMERIC_MEAN_FIELDS:
            source = valid_points if field in {
                "estimated_current_a",
                "current_sigma_a",
                "left_frequency_hz",
                "right_frequency_hz",
                "splitting_hz",
                "delta_f_sigma_hz",
                "common_mode_hz",
            } else points
            values = [
                value
                for point in source
                if (value := _finite_float(point.get(field))) is not None
            ]
            means[field] = _mean(values)

        timing_rates = [
            value
            for point in points
            if (
                value := _finite_float(
                    (point.get("timing") or {}).get("measured_update_rate_hz")
                )
            )
            is not None
        ]
        reasons = sorted(
            {
                str(point.get("invalid_reason")).strip()
                for point in points
                if point.get("invalid_reason")
            }
        )
        last = points[-1]
        elapsed_s = _finite_float(last.get("elapsed_s"))
        timestamp = (
            self.started_datetime + timedelta(seconds=elapsed_s)
            if elapsed_s is not None
            else datetime.now().astimezone()
        ).isoformat(timespec="milliseconds")
        return {
            "timestamp_local": timestamp,
            "elapsed_s": elapsed_s,
            "interval_s": float(interval_s),
            "current_a": means["estimated_current_a"],
            "current_std_a": _sample_std(current_values),
            "current_sigma_a": means["current_sigma_a"],
            "f_left_hz": means["left_frequency_hz"],
            "f_right_hz": means["right_frequency_hz"],
            "delta_f_hz": means["splitting_hz"],
            "delta_f_sigma_hz": means["delta_f_sigma_hz"],
            "common_mode_hz": means["common_mode_hz"],
            "valid_fraction": len(valid_points) / len(points),
            "all_samples_valid": len(valid_points) == len(points),
            "sample_count": len(points),
            "valid_sample_count": len(valid_points),
            "left_state": str(last.get("left_state") or ""),
            "right_state": str(last.get("right_state") or ""),
            "invalid_reason": ";".join(reasons),
            "relock_count": max(
                int(point.get("relock_count", 0) or 0) for point in points
            ),
            "lost_lock_count": max(
                int(point.get("lost_lock_count", 0) or 0) for point in points
            ),
            "left_error_hz": means["left_error_hz"],
            "right_error_hz": means["right_error_hz"],
            "left_quality": means["left_quality"],
            "right_quality": means["right_quality"],
            "measured_update_rate_hz": _mean(timing_rates),
        }

    def _write_row(self, row: dict[str, Any]) -> None:
        self._writer.writerow(
            {
                key: (
                    "true"
                    if row.get(key) is True
                    else "false"
                    if row.get(key) is False
                    else ""
                    if row.get(key) is None
                    else row.get(key)
                )
                for key in DATA_COLUMNS
            }
        )
        self._file.flush()
        self.rows_written += 1
        if bool(row.get("all_samples_valid")):
            self.valid_rows += 1
        self.last_saved_at = str(row["timestamp_local"])
        self.last_elapsed_s = _finite_float(row.get("elapsed_s"))
        if self.rows_written % 60 == 0:
            try:
                os.fsync(self._file.fileno())
            except OSError:
                pass
        self._write_metadata()

    def add_point(self, point: dict[str, Any]) -> bool:
        elapsed_s = _finite_float(point.get("elapsed_s"))
        if elapsed_s is None:
            return False
        with self.lock:
            if self.closed:
                return False
            if self.next_deadline_s is None:
                self.next_deadline_s = elapsed_s + self.interval_s
            if self.pending_points:
                previous_elapsed_s = _finite_float(
                    self.pending_points[-1].get("elapsed_s")
                )
                if (
                    previous_elapsed_s is not None
                    and elapsed_s - previous_elapsed_s > self.interval_s
                ):
                    self._write_row(
                        self._aggregate(self.pending_points, self.interval_s)
                    )
                    self.pending_points = [dict(point)]
                    self.next_deadline_s = elapsed_s + self.interval_s
                    return True
            self.pending_points.append(dict(point))
            if elapsed_s < self.next_deadline_s:
                return False
            row = self._aggregate(self.pending_points, self.interval_s)
            self._write_row(row)
            self.pending_points.clear()
            while self.next_deadline_s <= elapsed_s:
                self.next_deadline_s += self.interval_s
            return True

    def finalize(
        self,
        status: str,
        *,
        error_message: str = "",
        failed_stage: str = "",
        error_code: str = "",
        error_hint: str = "",
    ) -> dict[str, Any]:
        with self.lock:
            if not self.closed:
                if self.pending_points:
                    self._write_row(
                        self._aggregate(self.pending_points, self.interval_s)
                    )
                    self.pending_points.clear()
                self.status = str(status)
                if error_message:
                    self.error_message = str(error_message)
                if failed_stage:
                    self.failed_stage = str(failed_stage)
                if error_code:
                    self.error_code = str(error_code)
                if error_hint:
                    self.error_hint = str(error_hint)
                self.ended_at = datetime.now().astimezone().isoformat(
                    timespec="milliseconds"
                )
                try:
                    os.fsync(self._file.fileno())
                except OSError:
                    pass
                self._file.close()
                self.closed = True
                self._write_metadata()
        if self.rows_written:
            try:
                self.export_xlsx()
            except Exception as exc:
                with self.lock:
                    self.last_export_error = str(exc)
                    self._write_metadata()
        return self.status_dict()

    def _snapshot_csv(self) -> Path:
        with self.lock:
            if not self.closed:
                self._file.flush()
            snapshot = self.session_dir / f".snapshot_{uuid.uuid4().hex}.csv"
            shutil.copyfile(self.csv_path, snapshot)
            return snapshot

    def export_xlsx(self) -> Path:
        if Workbook is None or WriteOnlyCell is None:
            raise RuntimeError(
                "缺少 openpyxl，无法生成 Excel；CSV 恢复文件仍然安全保留。"
            )
        with self.export_lock:
            snapshot = self._snapshot_csv()
            temporary_xlsx = self.xlsx_path.with_suffix(".xlsx.tmp")
            try:
                self._build_workbook(snapshot, temporary_xlsx)
                os.replace(temporary_xlsx, self.xlsx_path)
                with self.lock:
                    self.last_export_error = ""
                    self._write_metadata()
                return self.xlsx_path
            finally:
                snapshot.unlink(missing_ok=True)
                temporary_xlsx.unlink(missing_ok=True)

    def _build_workbook(self, csv_path: Path, output_path: Path) -> None:
        workbook = Workbook(write_only=True)
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

        summary = workbook.create_sheet("Summary")
        data = workbook.create_sheet("Data")
        parameters = workbook.create_sheet("Parameters")

        for sheet in (summary, data, parameters):
            sheet.sheet_view.showGridLines = False

        dark_fill = PatternFill("solid", fgColor="123B5D")
        section_fill = PatternFill("solid", fgColor="DCEAF5")
        white_bold = Font(color="FFFFFF", bold=True)
        bold = Font(bold=True, color="16324F")
        light_border = Border(
            bottom=Side(style="thin", color="B7C9D6")
        )

        def styled_cell(
            sheet: Any,
            value: Any,
            *,
            fill: Any | None = None,
            font: Any | None = None,
            number_format: str | None = None,
            alignment: Any | None = None,
            border: Any | None = None,
        ) -> Any:
            cell = WriteOnlyCell(sheet, value=value)
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if number_format:
                cell.number_format = number_format
            if alignment is not None:
                cell.alignment = alignment
            if border is not None:
                cell.border = border
            return cell

        summary.freeze_panes = "A3"
        summary.column_dimensions["A"].width = 29
        summary.column_dimensions["B"].width = 30
        summary.column_dimensions["C"].width = 12
        summary.column_dimensions["D"].width = 52
        summary.append(
            [
                styled_cell(summary, "ODMR Dual-Peak Current Tracking", fill=dark_fill, font=white_bold),
                styled_cell(summary, "", fill=dark_fill),
                styled_cell(summary, "", fill=dark_fill),
                styled_cell(summary, "", fill=dark_fill),
            ]
        )
        summary_rows: list[tuple[str, Any, str, str]] = [
            ("Session ID", self.session_id, "", "Unique recording session"),
            ("Label", self.label, "", "User experiment label"),
            ("Status", self.status, "", "Recording terminal state"),
            ("Started", self.started_at, "", "Local time with timezone"),
            ("Ended", self.ended_at or "", "", "Local time with timezone"),
            ("Recording interval", self.interval_s, "s", "Aggregation interval"),
            ("Saved rows", self.rows_written, "rows", "Rows written to Data"),
            ("Valid rows", self.valid_rows, "rows", "Every source sample in interval was valid"),
        ]
        data_last_row = self.rows_written + 1
        if self.rows_written:
            summary_rows.extend(
                [
                    (
                        "Valid-row fraction",
                        f"=IFERROR(COUNTIF('Data'!M2:M{data_last_row},TRUE)/COUNT('Data'!B2:B{data_last_row}),0)",
                        "%",
                        "Fraction of fully valid one-second intervals",
                    ),
                    (
                        "Mean current",
                        f"=IFERROR(AVERAGE('Data'!D2:D{data_last_row}),\"\")",
                        "A",
                        "Mean of valid interval current values",
                    ),
                    (
                        "Current standard deviation",
                        f"=IFERROR(STDEV.S('Data'!D2:D{data_last_row}),\"\")",
                        "A",
                        "Long-run variation across saved intervals",
                    ),
                    (
                        "Minimum current",
                        f"=IFERROR(MIN('Data'!D2:D{data_last_row}),\"\")",
                        "A",
                        "Minimum valid interval current",
                    ),
                    (
                        "Maximum current",
                        f"=IFERROR(MAX('Data'!D2:D{data_last_row}),\"\")",
                        "A",
                        "Maximum valid interval current",
                    ),
                    (
                        "Mean splitting Δf",
                        f"=IFERROR(AVERAGE('Data'!I2:I{data_last_row}),\"\")",
                        "Hz",
                        "Mean physical resonance splitting",
                    ),
                ]
            )
        summary.append(
            [
                styled_cell(summary, "Metric", fill=section_fill, font=bold, border=light_border),
                styled_cell(summary, "Value", fill=section_fill, font=bold, border=light_border),
                styled_cell(summary, "Unit", fill=section_fill, font=bold, border=light_border),
                styled_cell(summary, "Definition", fill=section_fill, font=bold, border=light_border),
            ]
        )
        for label, value, unit, description in summary_rows:
            value_cell = styled_cell(summary, value)
            if isinstance(value, str) and value.startswith("="):
                value_cell.value = value
            if unit == "A":
                value_cell.number_format = "0.000000"
            elif unit == "Hz":
                value_cell.number_format = "0.000"
            elif unit == "s":
                value_cell.number_format = "0.000"
            elif unit == "%":
                value_cell.number_format = "0.00%"
            summary.append([label, value_cell, unit, description])
        data.freeze_panes = "A2"
        data.auto_filter.ref = f"A1:Y{max(1, data_last_row)}"
        widths = {
            "A": 25,
            "B": 14,
            "C": 18,
            "D": 16,
            "E": 16,
            "F": 20,
            "G": 22,
            "H": 22,
            "I": 20,
            "J": 24,
            "K": 25,
            "L": 15,
            "M": 17,
            "N": 18,
            "O": 23,
            "P": 19,
            "Q": 19,
            "R": 34,
        }
        for column, width in widths.items():
            data.column_dimensions[column].width = width
        for column in ("S", "T", "U", "V", "W", "X", "Y"):
            data.column_dimensions[column].width = 20
        data.append(
            [
                styled_cell(
                    data,
                    header,
                    fill=dark_fill,
                    font=white_bold,
                    alignment=Alignment(horizontal="center"),
                )
                for header in EXCEL_HEADERS
            ]
        )
        with csv_path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            for source_row in reader:
                excel_row: list[Any] = []
                for key in DATA_COLUMNS:
                    raw = source_row.get(key, "")
                    if key == "timestamp_local":
                        try:
                            value = datetime.fromisoformat(raw).replace(tzinfo=None)
                        except (TypeError, ValueError):
                            value = raw
                        excel_row.append(
                            styled_cell(
                                data,
                                value,
                                number_format="yyyy-mm-dd hh:mm:ss.000",
                            )
                        )
                    elif key in {"left_state", "right_state", "invalid_reason"}:
                        excel_row.append(raw)
                    elif key == "all_samples_valid":
                        excel_row.append(str(raw).lower() == "true")
                    elif key in {"sample_count", "valid_sample_count", "relock_count", "lost_lock_count"}:
                        excel_row.append(int(float(raw)) if raw else None)
                    else:
                        value = float(raw) if raw else None
                        number_format = (
                            "0.000000000"
                            if key in {"current_a", "current_std_a", "current_sigma_a"}
                            else "0.0000%"
                            if key == "valid_fraction"
                            else "#,##0.000"
                            if key
                            in {
                                "f_left_hz",
                                "f_right_hz",
                                "delta_f_hz",
                                "delta_f_sigma_hz",
                                "common_mode_hz",
                                "left_error_hz",
                                "right_error_hz",
                            }
                            else "0.0000"
                            if key in {"left_quality", "right_quality"}
                            else "0.000"
                        )
                        excel_row.append(
                            styled_cell(data, value, number_format=number_format)
                        )
                data.append(excel_row)

        parameters.freeze_panes = "A2"
        parameters.column_dimensions["A"].width = 48
        parameters.column_dimensions["B"].width = 68
        parameters.column_dimensions["C"].width = 20
        parameters.append(
            [
                styled_cell(parameters, "Parameter", fill=dark_fill, font=white_bold),
                styled_cell(parameters, "Value", fill=dark_fill, font=white_bold),
                styled_cell(parameters, "Source", fill=dark_fill, font=white_bold),
            ]
        )
        parameter_rows = _flatten_mapping(self.request_snapshot, "request")
        device_rows = _flatten_mapping(self.device_snapshot, "devices")
        for key, value in parameter_rows:
            parameters.append([key, value, "tracking request"])
        for key, value in device_rows:
            parameters.append([key, value, "device snapshot"])
        parameters.auto_filter.ref = (
            f"A1:C{1 + len(parameter_rows) + len(device_rows)}"
        )

        workbook.save(output_path)


class CurrentTrackingRecordingManager:
    def __init__(self, base_dir: Path) -> None:
        self.base_dir = Path(base_dir)
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self.lock = threading.RLock()
        self.active: CurrentTrackingRecorder | None = None
        self.sessions: dict[str, CurrentTrackingRecorder] = {}

    def start(
        self,
        *,
        interval_s: float,
        label: str,
        request_snapshot: dict[str, Any],
        device_snapshot: dict[str, Any],
    ) -> dict[str, Any]:
        with self.lock:
            if self.active is not None and not self.active.closed:
                self.active.finalize("superseded")
            recorder = CurrentTrackingRecorder(
                base_dir=self.base_dir,
                interval_s=interval_s,
                label=label,
                request_snapshot=request_snapshot,
                device_snapshot=device_snapshot,
            )
            self.active = recorder
            self.sessions[recorder.session_id] = recorder
            return recorder.status_dict()

    def record_point(self, point: dict[str, Any]) -> tuple[bool, dict[str, Any] | None]:
        with self.lock:
            recorder = self.active
        if recorder is None:
            return False, None
        wrote = recorder.add_point(point)
        return wrote, recorder.status_dict() if wrote else None

    def finish(
        self,
        status: str,
        *,
        error_message: str = "",
        failed_stage: str = "",
        error_code: str = "",
        error_hint: str = "",
    ) -> dict[str, Any] | None:
        with self.lock:
            recorder = self.active
            self.active = None
        return (
            recorder.finalize(
                status,
                error_message=error_message,
                failed_stage=failed_stage,
                error_code=error_code,
                error_hint=error_hint,
            )
            if recorder is not None
            else None
        )

    def status(self, session_id: str | None = None) -> dict[str, Any]:
        with self.lock:
            recorder = (
                self.sessions.get(session_id)
                if session_id
                else self.active
                or (next(reversed(self.sessions.values())) if self.sessions else None)
            )
        return recorder.status_dict() if recorder is not None else {
            "status": "idle",
            "session_id": None,
            "rows_written": 0,
            "download_available": False,
        }

    def export(self, session_id: str | None = None) -> tuple[Path, dict[str, Any]]:
        with self.lock:
            recorder = (
                self.sessions.get(session_id)
                if session_id
                else self.active
                or (next(reversed(self.sessions.values())) if self.sessions else None)
            )
        if recorder is None:
            raise FileNotFoundError("没有可导出的连续跟踪记录。")
        if recorder.rows_written <= 0:
            raise RuntimeError("记录尚未产生第一个聚合数据点。")
        path = recorder.export_xlsx()
        return path, recorder.status_dict()

